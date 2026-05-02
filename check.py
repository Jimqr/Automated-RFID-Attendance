import serial
import openpyxl
from datetime import datetime
import time
import os

# Change COM3 to your Arduino port
arduino = serial.Serial('COM3', 9600, timeout=1)
time.sleep(2)

# Load student database
db_file = "database.xlsx"
if not os.path.exists(db_file):
    print("Error: database.xlsx not found!")
    exit()

db_wb = openpyxl.load_workbook(db_file)
db_ws = db_wb["Students"]

# Attendance log
att_file = "attendance.xlsx"
if not os.path.exists(att_file):
    att_wb = openpyxl.Workbook()
    att_ws = att_wb.active
    att_ws.title = "Attendance"
    att_ws.append(["Name", "Grade & Section", "LRN", "UID", "Time"])
    att_wb.save(att_file)

def find_student(uid):
    for row in db_ws.iter_rows(min_row=2, values_only=True):
        db_uid, name, lrn, grade_section = row
        if db_uid == uid:
            return {"uid": db_uid, "name": name, "lrn": lrn, "grade_section": grade_section}
    return None

while True:
    if arduino.in_waiting > 0:
        uid = arduino.readline().decode().strip()
        if not uid:
            continue

        print("Card detected:", uid)
        student = find_student(uid)

        if student:  # Valid
            print("VALID:", student["name"])

            # Log attendance
            att_wb = openpyxl.load_workbook(att_file)
            att_ws = att_wb["Attendance"]
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            att_ws.append([student["name"], student["grade_section"], student["lrn"], uid, now])
            att_wb.save(att_file)

            # Tell Arduino it's valid
            arduino.write(b'VALID\n')

        else:  # Invalid
            print("INVALID UID")
            arduino.write(b'INVALID\n')